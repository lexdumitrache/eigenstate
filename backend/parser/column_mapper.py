"""CSV/Excel ingestion and column-mapping confirmation flow (spec §1, §4).

The LLM classifies each uploaded file into one of five roles (entity,
relationship, cost_matrix, calendar, parameter) and suggests column mappings.
The user must confirm entity and cost-matrix mappings before rows become
Entities; relationship and calendar tables are auto-confirmed and stored for
future modeling use.
"""
from __future__ import annotations

import csv
import io
import json

from spec.enums import TableRole
from spec.schema import (
    CalendarTable,
    ColumnMapping,
    Entity,
    PairwiseCostTable,
    RelationshipTable,
)
from .llm_adapter import LLMAdapter

CLASSIFY_PROMPT = """Classify this CSV table and propose how its columns map to the optimization model.

Table roles:
  "entity"       — each row is a standalone object: a van, package, employee, job, location
  "relationship" — each row links two entity types: driver→route, employee→shift, van→location
  "cost_matrix"  — each row gives a numeric cost/time/distance between two entity types
  "calendar"     — each row describes availability or scheduling windows for an entity
  "parameter"    — rows contain global configuration values (not individual entities)

Output ONLY valid JSON. The required fields depend on the table_role:

For "entity":
{"table_role": "entity", "entity_category": "<singular_snake_case>",
 "id_column": "<col or null>", "column_to_field": {"<csv_col>": "<field_name>"}}

For "cost_matrix":
{"table_role": "cost_matrix",
 "from_column": "<col>", "to_column": "<col>", "cost_column": "<col>",
 "from_category": "<category>", "to_category": "<category>"}

For "relationship":
{"table_role": "relationship",
 "from_column": "<col>", "to_column": "<col>",
 "from_category": "<category>", "to_category": "<category>",
 "value_column": "<col or null>",
 "column_to_field": {"<remaining_col>": "<field_name>"}}

For "calendar":
{"table_role": "calendar", "entity_column": "<col>", "entity_category": "<category>",
 "start_column": "<col or null>", "end_column": "<col or null>",
 "column_to_field": {"<remaining_col>": "<field_name>"}}

For "parameter":
{"table_role": "parameter", "column_to_field": {"<csv_col>": "<field_name>"}}

Rules:
- Use snake_case for all field names; embed units when unambiguous (e.g. "capacity_kg")
- entity_category / from_category / to_category must be singular snake_case: "van", "driver"
- from_category / to_category should match entity categories from related entity tables when possible
"""


def read_table(file_name: str, content: bytes,
               sheet_name: str | None = None) -> tuple[list[str], list[dict]]:
    """Return (header, rows) from CSV or Excel bytes."""
    if file_name.lower().endswith((".xlsx", ".xls")):
        try:
            import openpyxl
        except ImportError as e:
            raise ValueError("Excel support requires openpyxl; upload CSV instead.") from e
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        if sheet_name is not None:
            if sheet_name not in wb.sheetnames:
                raise ValueError(
                    f"Sheet '{sheet_name}' not found in '{file_name}'. "
                    f"Available sheets: {wb.sheetnames}"
                )
            ws = wb[sheet_name]
        else:
            ws = wb.worksheets[0]
        it = ws.iter_rows(values_only=True)
        header = [str(h) if h is not None else f"col_{j}" for j, h in enumerate(next(it))]
        rows = [dict(zip(header, r)) for r in it if any(v is not None for v in r)]
        return header, rows
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    rows = list(reader)
    return list(reader.fieldnames or []), rows


def classify_table(
    file_name: str, header: list[str], sample_rows: list[dict], adapter: LLMAdapter
) -> ColumnMapping | PairwiseCostTable | RelationshipTable | CalendarTable:
    """Ask the LLM to classify a file and return the appropriate typed mapping."""
    payload = json.dumps({"file": file_name, "header": header,
                          "sample_rows": sample_rows[:5]})
    raw = adapter.complete_json(CLASSIFY_PROMPT, payload)
    role = raw.get("table_role", "entity")

    if role == TableRole.COST_MATRIX:
        return PairwiseCostTable(
            file_name=file_name,
            agent_column=raw.get("from_column") or header[0],
            task_column=raw.get("to_column") or header[1],
            cost_column=raw.get("cost_column") or header[2] if len(header) > 2 else header[1],
            agent_category=raw.get("from_category") or "agent",
            task_category=raw.get("to_category") or "task",
            confirmed=False,
        )

    if role == TableRole.RELATIONSHIP:
        return RelationshipTable(
            file_name=file_name,
            from_column=raw.get("from_column") or header[0],
            to_column=raw.get("to_column") or header[1],
            from_category=raw.get("from_category") or "entity",
            to_category=raw.get("to_category") or "entity",
            value_column=raw.get("value_column"),
            column_to_field=raw.get("column_to_field") or {},
            confirmed=False,
        )

    if role == TableRole.CALENDAR:
        return CalendarTable(
            file_name=file_name,
            entity_column=raw.get("entity_column") or header[0],
            entity_category=raw.get("entity_category") or "entity",
            start_column=raw.get("start_column"),
            end_column=raw.get("end_column"),
            column_to_field=raw.get("column_to_field") or {},
            confirmed=False,
        )

    # entity or parameter both produce a ColumnMapping
    table_role = TableRole.PARAMETER if role == TableRole.PARAMETER else TableRole.ENTITY
    return ColumnMapping(
        file_name=file_name,
        entity_category=raw.get("entity_category") or "item",
        id_column=raw.get("id_column"),
        column_to_field=raw.get("column_to_field") or {c: c for c in header},
        table_role=table_role,
        confirmed=False,
    )


# ── backwards-compat helpers kept for tests / direct callers ─────────────────

def suggest_mapping(file_name: str, header: list[str], sample_rows: list[dict],
                    adapter: LLMAdapter) -> ColumnMapping:
    result = classify_table(file_name, header, sample_rows, adapter)
    if isinstance(result, ColumnMapping):
        return result
    # Fallback: treat as entity table
    return ColumnMapping(
        file_name=file_name,
        entity_category="item",
        id_column=None,
        column_to_field={c: c for c in header},
        confirmed=False,
    )


def suggest_pairwise_table(file_name: str, header: list[str], sample_rows: list[dict],
                           adapter: LLMAdapter) -> PairwiseCostTable | None:
    result = classify_table(file_name, header, sample_rows, adapter)
    return result if isinstance(result, PairwiseCostTable) else None


# ─────────────────────────────────────────────────────────────────────────────

def _coerce(value):
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    s = str(value).strip()
    try:
        f = float(s)
        return int(f) if f.is_integer() else f
    except ValueError:
        return s


def rows_to_entities(mapping: ColumnMapping, rows: list[dict]) -> list[Entity]:
    if not mapping.confirmed:
        raise ValueError("Column mapping must be confirmed before ingestion.")
    entities = []
    seen_ids: set[str] = set()
    for i, row in enumerate(rows):
        attrs = {field: _coerce(row.get(col))
                 for col, field in mapping.column_to_field.items()}
        if mapping.id_column and row.get(mapping.id_column) is not None:
            eid = str(row[mapping.id_column])
        else:
            eid = f"{mapping.entity_category}_{i + 1}"
        if eid in seen_ids:
            raise ValueError(
                f"Duplicate ID '{eid}' in '{mapping.file_name}' "
                f"(row {i + 1}). Each row must have a unique ID."
            )
        seen_ids.add(eid)
        entities.append(Entity(id=eid, category=mapping.entity_category,
                               attributes=attrs, source_file=mapping.file_name))
    return entities
